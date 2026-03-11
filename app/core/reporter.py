import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.models.db import PolicyRecord


def _format_numbered_lines(text: str) -> str:
    """在序号（1. 2、1） 等）前插入换行，改善 Excel 单元格可读性。"""
    if not text or text == "/":
        return text
    # 匹配：前一字符不是换行 + 序号前空白 + 序号（数字+点/顿号，点后非数字避免处理小数）
    return re.sub(r'(?<=[^\n])[ \t]*(\d+[.．、](?!\d))', r'\n\1', text).strip()

# 列顺序与列宽（字符数）集中管理，调整字段顺序只改这里
COLUMNS: list[tuple[str, int]] = [
    ("项目名称",   30),
    ("政策依据",   25),
    ("归口部门",   20),
    ("联系人",     28),
    ("申报时间",   22),
    ("支持方向",   30),
    ("特定方向要求", 32),
    ("申报要求",   40),
    ("优惠政策",   30),
    ("申报材料",   35),
    ("申报方式",   30),
    ("网站链接",   32),
    ("政策有效期", 20),
]

# 前 5 列：按「同一政策文档」分组合并
_GROUP_COL_COUNT = 5

# 第 7-13 列（跳过第 6 列「支持方向」）：组内相邻相同值合并
_VALUE_MERGE_COL_START = 7   # 1-indexed，含
_VALUE_MERGE_COL_END   = len(COLUMNS)  # 1-indexed，含

_HEADER_FILL = PatternFill(fill_type="solid", fgColor="2E75B6")
_HEADER_FONT = Font(bold=True, color="FFFFFF")

_ALIGN_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="left")
_ALIGN_TOP    = Alignment(wrap_text=True, vertical="top")
_LINK_FONT    = Font(color="0563C1", underline="single")  # Excel 超链接默认样式

_LINK_COL_NAME = "网站链接"


def export_to_excel(task_id: str, records: list[PolicyRecord], output_dir: Path) -> Path:
    """将本次任务的 PolicyRecord 列表导出为 Excel 文件，返回生成文件的路径。"""
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "政策摘要"

    # ── 表头 ──
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = col_width
    ws.row_dimensions[1].height = 30

    # ── 数据行 ──
    for row_idx, record in enumerate(records, start=2):
        for col_idx, (col_name, _) in enumerate(COLUMNS, start=1):
            value = getattr(record, col_name, "/") or "/"
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_name == _LINK_COL_NAME and value != "/":
                # 写入超链接，让 Excel 可直接点击跳转
                cell.value = value
                cell.hyperlink = value
                cell.font = _LINK_FONT
                cell.alignment = _ALIGN_TOP
            else:
                cell.value = _format_numbered_lines(value)
                cell.alignment = _ALIGN_TOP

    # ── 合并单元格 ──
    if len(records) >= 2:
        groups = _build_groups(records)
        _merge_group_cols(ws, groups)           # 前 5 列：按分组合并
        _merge_value_cols(ws, records, groups)  # 第 7-13 列：组内相同值合并

    output_path = output_dir / f"{task_id}.xlsx"
    wb.save(output_path)
    return output_path


# ── 辅助：按前 5 列内容划分分组区间 ──────────────────────────────────────

def _build_groups(records: list) -> list[tuple[int, int]]:
    """
    按前 5 列的值将 records 划分为连续分组，返回 (start_idx, end_idx) 列表。
    start_idx / end_idx 均为 records 的 0-based 下标。
    """
    col_names = [name for name, _ in COLUMNS[:_GROUP_COL_COUNT]]

    def _key(r):
        return tuple(getattr(r, n, "/") or "/" for n in col_names)

    groups: list[tuple[int, int]] = []
    group_start = 0
    for i in range(1, len(records)):
        if _key(records[i]) != _key(records[group_start]):
            groups.append((group_start, i - 1))
            group_start = i
    groups.append((group_start, len(records) - 1))
    return groups


# ── 合并：前 5 列按分组整体合并 ─────────────────────────────────────────

def _merge_group_cols(ws, groups: list[tuple[int, int]]) -> None:
    """对每个跨多行的分组，将前 5 列全部合并。"""
    for start_idx, end_idx in groups:
        if start_idx == end_idx:
            continue
        start_row = start_idx + 2   # Excel 行号（1=表头，2=首数据行）
        end_row   = end_idx   + 2
        for col in range(1, _GROUP_COL_COUNT + 1):
            ws.merge_cells(
                start_row=start_row, start_column=col,
                end_row=end_row,   end_column=col,
            )
            ws.cell(row=start_row, column=col).alignment = _ALIGN_CENTER


# ── 合并：第 7-13 列在组内按相同值合并 ───────────────────────────────────

def _merge_value_cols(ws, records: list, groups: list[tuple[int, int]]) -> None:
    """
    在每个分组内，对第 7-13 列逐列扫描：
    相邻行中值完全相同（包括 "/"）则合并，值不同则断开。
    """
    for start_idx, end_idx in groups:
        if start_idx == end_idx:
            continue  # 单行分组，无需合并
        for col_idx in range(_VALUE_MERGE_COL_START, _VALUE_MERGE_COL_END + 1):
            _merge_col_within_range(ws, records, start_idx, end_idx, col_idx)


def _merge_col_within_range(
    ws, records: list, start_idx: int, end_idx: int, col_idx: int
) -> None:
    """
    在 records[start_idx..end_idx] 范围内，对 col_idx 列扫描连续相同值并合并。
    Excel 行号 = records 下标 + 2。
    """
    col_name   = COLUMNS[col_idx - 1][0]
    seg_start  = start_idx  # 当前连续相同值段的起始下标

    def _val(idx: int) -> str:
        return getattr(records[idx], col_name, "/") or "/"

    for i in range(start_idx + 1, end_idx + 1):
        if _val(i) != _val(i - 1):
            # 值发生变化，尝试合并 [seg_start, i-1]
            if seg_start < i - 1:
                _do_merge(ws, seg_start + 2, i - 1 + 2, col_idx)
            seg_start = i

    # 处理末尾段 [seg_start, end_idx]
    if seg_start < end_idx:
        _do_merge(ws, seg_start + 2, end_idx + 2, col_idx)


def _do_merge(ws, start_row: int, end_row: int, col_idx: int) -> None:
    ws.merge_cells(
        start_row=start_row, start_column=col_idx,
        end_row=end_row,   end_column=col_idx,
    )
    ws.cell(row=start_row, column=col_idx).alignment = _ALIGN_CENTER
